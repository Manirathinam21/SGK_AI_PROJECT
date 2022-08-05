#-------------------------------------------- Fontem_Excel ------------------------------------------------------------------#
#Packages used
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from langid import classify
import string
import re
import joblib
#--------------------------------------------------------------------------------------------------------------------------------
#Laser Embedding

from laserembeddings import Laser
path_to_bpe_codes = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fcodes'
path_to_bpe_vocab = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fvocab'
path_to_encoder = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'
laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder) 
# --------------------------------------------------------------------------------------------------------------------------------
# Loading MLP Classifier for classification keys

classifier=joblib.load('/Users/manirathinams/Documents/Python /Fontem Excel/Fontem_excel.sav')
# ----------------------------------------------------------------------------------------------------------------------------------------------------------------
## To bring a dataframe with bold char & percentage values in percentile cells along with coordinates

def Extracting_coordinates(file, sheet_name):
    wb=load_workbook(file)
    sheet=wb.sheetnames
    ws=wb[sheet_name]
    #ws=wb[sheet[2]]
    row=ws.max_row
    col=ws.max_column

    Newlist=[]
    for i in range(1, row+1):
        temp=[]
        for j in range(1, col+1):
            if (ws.cell(i,j).font.bold==True) and (ws.cell(i,j).value!=None):
                temp.append([ws.cell(i,j).coordinate, '&lt;b&gt;'+ws.cell(i,j).value+'&lt;/b&gt;'])
            elif (ws.cell(i,j).number_format in ['0.0%','0.00%','0.000%','0.0000%']) and ('MG/ML' or 'mg/ml' in str(ws.cell(i,j).value)):
                temp.append([ws.cell(i,j).coordinate, ws.cell(i,j).value])
            elif (ws.cell(i,j).number_format in ['0.0%', '0.00%', '0.000%','0.0000%']) and (ws.cell(i,j).value!=None):
                temp.append([ws.cell(i,j).coordinate, str(ws.cell(i,j).value*100)+'%'])
            else:
                temp.append([ws.cell(i,j).coordinate, ws.cell(i,j).value])
        Newlist.append(temp)

    gen=pd.DataFrame(Newlist)
    return gen
# --------------------------------------------------------------------------------------------------------
#general information slicing function

def slicing_content(gen):
    r_slice, c_slice=[], []
    for k in range(len(gen)):
        for h in gen:
            if 'blu product' in str(gen.iloc[k,h][1]).lower():
                r_slice.append(k) 
                c_slice.append(h)
    ##To extract muultiple tables with 'blu product' as header in same sheet            
    all_tables=[]            
    for i in range(len(r_slice)):
        try:
            all_tables.append(gen.iloc[r_slice[i]:r_slice[i+1],c_slice[i]:])
        except:
            all_tables.append(gen.iloc[r_slice[i]:,c_slice[i]:])

    # Removing Null cells
    dataframes=[]
    for data in all_tables:
        clean=[]
        for i in range(len(data.columns)):
            c=[]
            for j in range(len(data)):
                if data.iloc[j,i]!=None:
                    c.append(data.iloc[j,i])
            if c:
                clean.append(c) 
        dataframes.append(pd.DataFrame(clean))
    return dataframes
# --------------------------------------------------------------------------------------------------------------------
# key_val classification

def Key_val(dataframes):
    key=[]
    val=[]
    for df in dataframes:
        for s in range(len(df)):
            if df.iloc[s,0]!=None and df.iloc[s,0][1]!=None:
                x=(df.iloc[s,0][1].replace('&lt;b&gt;','').replace('&lt;/b&gt;',''))
                ## Removing the elements within brackets & list along ()braces and removing digits
                head=(re.sub("[\(\[].*?[\)\]]","", x).replace('\n','').strip())
                ky=(re.sub(r"\d","",head))
                ky_cln=ky.replace("CLP Article","").replace("EUTPD Article","").replace('BACK','').replace('FRONT','').replace('©','').replace('.','').replace('SIDE PANEL','').replace('BOTTOM','').replace('FRONT & TOP','').replace('RIGHT AND LEFT SIDE','').replace('RIGHT & LEFT SIDE, TOP','').strip()
                proba=classifier.predict_proba(laser.embed_sentences(ky_cln, lang='en'))[0]
                proba.sort()
                content=[]
                for t in range(1, len(df.columns)):
                    if df.iloc[s,t][1]!=None:
                        content.append({str(df.iloc[s,t][0])+'_'+str(classify(str(df.iloc[s,t][1]))[0]):str(df.iloc[s,t][1]).replace('\n', '').strip()})
                if content:
                    val.append(content) 
                    if proba[-1] >0.65:
                        key.append((classifier.predict(laser.embed_sentences(ky_cln, lang='en')))[0])
                    else:
                        key.append('UNMAPPED')
    return key,val
# --------------------------------------------------------------------------------------------------------------------
# Removing unwanted values 

def val_clean(val_temp):
    remove=[]
    for i in range(len(val_temp)):
        temp=val_temp[i]
        for j in range(len(temp)):
            if list(temp[j].values())[0] !=None:
                if ('MFP INTENSE STARTER KIT').lower() in str(list(temp[j].values())[0]).lower(): 
                    remove.append([i,j])
                elif ('Key Product Reference Information').lower() in str(list(temp[j].values())[0]).lower():
                    remove.append([i,j])
                elif ('Front Panel').lower() in str(list(temp[j].values())[0]).lower():
                    remove.append([i,j])
                elif ('Left Panel').lower() in str(list(temp[j].values())[0]).lower():
                    remove.append([i,j])
                elif ('Right Panel').lower() in str(list(temp[j].values())[0]).lower():
                    remove.append([i,j])
                elif ('Back Panel').lower() in str(list(temp[j].values())[0]).lower():
                    remove.append([i,j])
            else:
                remove.append([i,j])

    remove.sort(reverse=True)  

    for i in range(len(remove)):
        val_temp[remove[i][0]].remove(val_temp[remove[i][0]][remove[i][1]])
    return val_temp
# -----------------------------------------------------------------------------------------------------------------------
#Fontem_excel main function

def fontem_main(file, sheet_name):
    gen=Extracting_coordinates(file, sheet_name)
    dataframes=slicing_content(gen)
    key,val_temp=Key_val(dataframes)
    val=val_clean(val_temp)
    general={}
    if len(key)==len(val):
        for m in range(len(key)):
            general.setdefault(key[m],[]).extend(val[m])
    return general