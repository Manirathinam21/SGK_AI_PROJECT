#-------------------------------------------- Kelloggs_Excel ------------------------------------------------------------------#
#Packages used
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from langid import classify
import string
import re
import joblib
#--------------------------------------------------------------------------------------------------------------------------------
#Laser Embedding

from laserembeddings import Laser
path_to_bpe_codes = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fcodes'
path_to_bpe_vocab = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/93langs.fvocab'
path_to_encoder = r'/Users/manirathinams/opt/anaconda3/lib/python3.9/site-packages/laserembeddings/data/bilstm.93langs.2018-12-26.pt'
laser = Laser(path_to_bpe_codes, path_to_bpe_vocab, path_to_encoder) 
# --------------------------------------------------------------------------------------------------------------------------------
# Loading MLP Classifier for classification keys

classifier=joblib.load('/Users/manirathinams/Documents/KT/Kelloggs_Excel/kelloggs_model.sav')
#---------------------------------------------------------------------------------------------------------------------------------
## To bring a dataframe with bold char & percentage values in percentile cells along with coordinates

def Extracting_coordinates(file, sheet_name):
    wb=load_workbook(file)
    sheet=wb.sheetnames
    ws=wb[sheet_name]
    #ws=wb[sheet[2]]
    row=ws.max_row
    col=ws.max_column

    Newlist=[]
    for i in range(1,row+1):
        temp=[]
        for j in range(1,col+1):
            ## Getting cell values & coordinate's along the values which are in bold char to specific format,removing null cell's which is bold
            if (ws.cell(i,j).font.bold==True) and (ws.cell(i,j).value!=None):
                    temp.append([ws.cell(i,j).coordinate, '&lt;b&gt;'+ws.cell(i,j).value+ '&lt;/b&gt;'])
            ## Converting the decimal values in Nutrition table percentage column to percentage value
            elif ws.cell(i,j).number_format in ['0%','0.0%','0.00%', '0.000%']:
                temp.append([ws.cell(i,j).coordinate, str(ws.cell(i,j).value*100)+"%"])
            else:
                 temp.append([ws.cell(i,j).coordinate, ws.cell(i,j).value])
        Newlist.append(temp)

    gen=pd.DataFrame(Newlist)
    return gen
#-----------------------------------------------------------------------------------------------------------------------------------------
##To get a general information of master copy sheet from date to RMS supplier code

def general_information_slicing(gen):
    ind=[]
    for i in range(len(gen)):
        if 'date' in str(gen.iloc[i,0][1]).lower().strip():
            ind.append(i)
        elif 'item' in str(gen.iloc[i,0][1]).lower().strip():
            ind.append(i)
            break
    s,t=ind

    dataf=gen.iloc[s:t,:].reset_index(drop=True)
    ## Removing Null columns
    tem=[]
    for p in range(len(dataf)):
        r=[]
        for q in range(len(dataf.columns)):
            if dataf.iloc[p,q][1]!=None:
                r.append(dataf.iloc[p,q])
        if r:
            tem.append(r)

    ## To get a general information, filetring a data btwn front panel & Nutrition information
    index=[]
    for i in range(len(gen)):
        if 'front panel' in str(gen.iloc[i,0][1]).lower().strip():
            index.append(i)
        elif 'nutrition information' in str(gen.iloc[i,0][1]).lower().strip():
            index.append(i)
            break
    a,b=index

    df=gen.iloc[a:b,:].reset_index(drop=True)

    ## Removing 'B' columns and Null columns
    data=[]
    for i in range(len(df)):
        r=[]
        for j in df:
            if df.iloc[i,j][0][0]!='B' and df.iloc[i,j][0][0]!='G' and df.iloc[i,j][1]!=None: 
                r.append(df.iloc[i,j])
        if r:
            data.append(r)

    merge=tem+data
    df1=pd.DataFrame(merge)
    return df1
#---------------------------------------------------------------------------------------------------------------------------------------------
##Function to get a json response of general information

def gen_information(df1):
    item=[]
    content=[]
    for i in range(len(df1)):
        if df1.iloc[i,1]!=None:
        ##Getting detials in seperate dictionary with coordinates,language classification as a key & information as a value 
            content.append({str(df1.iloc[i,1][0])+'_'+str(classify(str(df1.iloc[i,1][1]))[0]):str(df1.iloc[i,1][1]).replace('<','&lt;').replace('>','&gt;').replace('\n','').strip()})
        ##Getting items in seperate list along removing :,# in tiems
            x=(df1.iloc[i,0][1].replace(':','').replace('#',''))
        ## Removing the elements within brackets & list along () --eg: ('expiration date')& ['recycleable'] in same or next line also removes '\n'
            key=(re.sub("[\(\[].*?[\)\]]","", x).replace('\n','').strip())
        ## It w'll give the probability of occuring value for each target class element and we sort the prob value in ascending order
            prob=classifier.predict_proba(laser.embed_sentences(key, lang='en'))[0]
            prob.sort()
        ## we get the max probability occuring value and check whether it's above 0.65% of chance of occuring particular target else we choose the target variable as unmapped 
            if prob[-1]>0.65:
        ##Converting the key to vector format through laser & then predict the target variable for that particular vector value using already trained model 'Classifier'
                item.append((classifier.predict(laser.embed_sentences(key, lang='en')))[0])
            else:
                item.append('UNMAPPED')

    ##create a dictionary with items as a key & detail contents as a value        
    gen_infor={}
    for i in range(len(item)):
    ##setdefault will allow duplicate keys in dictionary
        gen_infor.setdefault(item[i], []).append(content[i])
        
    return gen_infor
#--------------------------------------------------------------------------------------------------------------------
#Nutrition table function

def Nutrition(gen):
    ## Getting index of multiple nutrition tables
    header=[]
    footer=[]
    for i in range(len(gen)):
        if 'nutrition information' in str(gen.iloc[i,0][1]).lower().strip():
            header.append(i)
        elif 'daily intakes' in str(gen.iloc[i,0][1]).lower().strip():
            footer.append(i)
        
    final=[]
    ##writing a loop for executing a code no of times to get json as equal to no of nutrition tables
    if len(header)== len(footer):
        for w in range(len(header)):   
            nut=gen.iloc[header[w]:footer[w],:].reset_index(drop=True)

            ## Removing Null columns
            tabls=[]
            dat=[]
            for i in range(len(nut)):
                r=[]
                for k in nut:
                    if nut.iloc[i,k][1]!=None: 
                        r.append(nut.iloc[i,k])
                if r:
                    dat.append(r)
            dfs=pd.DataFrame(dat)
            dfs = dfs.replace(to_replace='None', value=np.nan).dropna()
            dfs=dfs.reset_index(drop=True)
            ##Extracting a columns with percentage as a PDV & normal strings as Values along with language classification  
            value=[]
            for j in range(1,len(dfs.columns)):
                temp=[]
                for i in range(len(dfs)):
                    if '%' in str(dfs.iloc[i,j][1]):
                        temp.append({'PDV':{str(dfs.iloc[i,j][0])+'_'+str(classify(str(dfs.iloc[i,j][1]))[0]):dfs.iloc[i,j][1]}})
                    else:
                        temp.append({'Value':{str(dfs.iloc[i,j][0])+'_'+str(classify(str(dfs.iloc[i,j][1]))[0]):dfs.iloc[i,j][1]}})
                value.append(temp)

            ##Extracting nutrition header columns seperately like Energy,protein,carbohydrates
            head=[]
            for i in range(len(dfs)):
                key=(dfs.iloc[i,0][1].replace('-','').replace(',','').strip())
                proba=classifier.predict_proba(laser.embed_sentences(key, lang='en'))[0]
                proba.sort()
                if proba[-1] >0.65:
                    head.append((classifier.predict(laser.embed_sentences(key, lang='en')))[0])
                else:
                    head.append('UNMAPPED')
                    
            ##Mapping headers with Nutrition value contents    
            final.append({h:list(v) for h,v in zip(head,zip(*value))})
        return final
#------------------------------------------------------------------------------------------------------------------------
#Kelloggs main Function

def kelloggs_main(file, sheet_name):
    gen=Extracting_coordinates(file, sheet_name)
    dataf=general_information_slicing(gen)
    general=gen_information(dataf)
    general['NUTRITION_FACTS']=Nutrition(gen)
    return general

